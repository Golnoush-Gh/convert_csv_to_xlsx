#!/usr/bin/python3

import os
import locale
import csv
import click
import openpyxl

default_encoding = locale.getpreferredencoding()

def convert_csv_to_xlsx(csv_filenames, xlsx_filename, csv_delimiter=",", csv_file_encoding=default_encoding):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    
    for csv_filename in csv_filenames:
        name, ext = os.path.splitext(csv_filename)
        worksheet = workbook.create_sheet(name)        
        with open(csv_filename, "r", encoding=csv_file_encoding) as fobj:
            csv_reader = csv.reader(fobj, delimiter=csv_delimiter)
            for row_index, row in enumerate(csv_reader):
                for col_index, value in enumerate(row):
                    worksheet.cell(row_index + 1, col_index + 1).value = value

    workbook.save(xlsx_filename)


@click.command()
@click.option("-d", "--delimiter", default=",", help="""CSV column delimiter, default to ",".""")
@click.option("-e", "--encoding", default=default_encoding, help="CSV file encoding, default to {0}.".format(default_encoding))
@click.argument("csv_file", nargs=-1, required=True)
@click.argument("xlsx_file", nargs=1, required=True)
def do_convert(delimiter, encoding, csv_file, xlsx_file):
    """Convert csv file to xlsx file.

    Usage:

    convert-multi-csv-to-xlsx test1.csv test2.csv ... bundled.xlsx
    """
    convert_csv_to_xlsx(csv_file, xlsx_file, delimiter, encoding)
    print("Excel file created at: {0}".format(xlsx_file))

if __name__ == "__main__":
    do_convert()
